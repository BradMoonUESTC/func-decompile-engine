import os
import re
import json
import csv
import openpyxl
import os
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import re
def find_fif_functions(text, filename, hash):
    regex = r"([\w$]+(?:[$][\w$]+)*)\s+([A-Z]+):<\{"
    matches = re.finditer(regex, text)

    functions = []
    lines = text.split('\n')
    line_starts = {i: sum(len(line) + 1 for line in lines[:i]) for i in range(len(lines))}

    # First, collect all function bodies to construct the complete contract code
    function_bodies = []
    for match in matches:
        angle_brace_count = 1
        function_body_start = match.start()
        inside_angle_braces = True

        for i in range(match.end(), len(text)):
            if text[i:i+2] == '<{':
                angle_brace_count += 1
            elif text[i:i+2] == '}>':
                angle_brace_count -= 1

            if inside_angle_braces and angle_brace_count == 0:
                function_body_end = i + 2
                function_bodies.append(text[function_body_start:function_body_end])
                break

    # Complete contract code string
    contract_code = "\n".join(function_bodies).strip()

    # Iterate through matches again to create function definitions
    for match in re.finditer(regex, text):
        start_line_number = next(i for i, pos in line_starts.items() if pos > match.start()) - 1
        function_header = match.group(0)
        
        angle_brace_count = 1
        function_body_start = match.start()
        inside_angle_braces = True

        for i in range(match.end(), len(text)):
            if text[i:i+2] == '<{':
                angle_brace_count += 1
            elif text[i:i+2] == '}>':
                angle_brace_count -= 1

            if inside_angle_braces and angle_brace_count == 0:
                function_body_end = i + 2
                end_line_number = next(i for i, pos in line_starts.items() if pos > function_body_end) - 1
                function_body = text[function_body_start:function_body_end]
                function_body_lines = function_body.count('\n') + 1
                break

        # Extract function name
        func_name = match.group(1)
        
        # Extract modifier
        modifier = match.group(2)

        functions.append({
            'type': 'FunctionDefinition',
            'name': func_name,
            'start_line': start_line_number + 1,
            'end_line': end_line_number,
            'offset_start': 0,
            'offset_end': 0,
            'content': function_body,
            'contract_name': filename.replace('.code.fif', ''),
            'contract_code': contract_code,
            'modifiers': [modifier],
            'stateMutability': None,
            'returnParameters': None,
            'visibility': 'public',  # Assuming all functions are public in .fif files
            'node_count': function_body_lines
        })

    return functions
def process_fif_file(file_path, hash_value):
    with open(file_path, 'r', encoding='utf-8') as file:
        fif_code = file.read()
    
    filename = os.path.basename(file_path)
    functions = find_fif_functions(fif_code, filename, hash_value)
    
    return functions

def process_fif_folder(folder_path, hash_value):
    all_functions = []
    
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith('.fif'):
                file_path = os.path.join(root, file)
                print(f"Processing file: {file_path}")
                functions = process_fif_file(file_path, hash_value)
                all_functions.extend(functions)
                
                print(f"\nParsing results for {file_path}:")
                for func in functions:
                    pass
                    # print(json.dumps(func, indent=2))
                    # print("-" * 50)
                
                print(f"Found {len(functions)} functions in {file_path}.")
    
    return all_functions

def test_find_fif_functions(folder_path, hash_value, output_excel):
    if not os.path.isdir(folder_path):
        print(f"Error: '{folder_path}' is not a valid folder path.")
        return

    print(f"Processing folder: {folder_path}")
    all_functions = process_fif_folder(folder_path, hash_value)

    print(f"\nTotal functions found: {len(all_functions)}.")

    # 定义要排除的函数名列表
    excluded_names = {'EQINT', 'EQUAL', 'GTINT', 'ISNULL', 'LESS', 'NEQINT', 'NOT', 'PUSH', 'SWAP', 'XCHG'}

    # 过滤函数
    filtered_functions = [func for func in all_functions if func['name'] not in excluded_names]

    print(f"Functions after filtering: {len(filtered_functions)}")

    # Write results to Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FIF Functions"

    headers = ['Name', 'Start Line', 'End Line', 'Content', 'Contract Name', 'Contract Code', 'Modifiers', 'Visibility', 'Node Count']
    ws.append(headers)

    # Style the header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for func in filtered_functions:
        ws.append([
            func['name'],
            func['start_line'],
            func['end_line'],
            func['content'],
            func['contract_name'],
            func['contract_code'],
            ','.join(func['modifiers']),
            func['visibility'],
            func['node_count']
        ])

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_excel)

    print(f"Results have been written to {output_excel}")

    return filtered_functions
def run_fif_decomposer(folder_path, hash_value, output_excel):
    if not os.path.isdir(folder_path):
        print(f"Error: '{folder_path}' is not a valid folder path.")
        return None

    print(f"Processing folder: {folder_path}")
    all_functions = process_fif_folder(folder_path, hash_value)

    print(f"\nTotal functions found: {len(all_functions)}.")

    # 定义要排除的函数名列表
    excluded_names = {'EQINT', 'EQUAL', 'GTINT', 'ISNULL', 'LESS', 'NEQINT', 'NOT', 'PUSH', 'SWAP', 'XCHG'}

    # 过滤函数
    filtered_functions = [func for func in all_functions if func['name'] not in excluded_names]

    print(f"Functions after filtering: {len(filtered_functions)}")

    # Write results to Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FIF Functions"

    headers = ['Name', 'Start Line', 'End Line', 'Content', 'Contract Name', 'Contract Code', 'Modifiers', 'Visibility', 'Node Count']
    ws.append(headers)

    # Style the header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for func in filtered_functions:
        ws.append([
            func['name'],
            func['start_line'],
            func['end_line'],
            func['content'],
            func['contract_name'],
            func['contract_code'],
            ','.join(func['modifiers']),
            func['visibility'],
            func['node_count']
        ])

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_excel)

    print(f"Results have been written to {output_excel}")

    return filtered_functions

if __name__ == "__main__":
    FOLDER_PATH = "test/DexRouter"  # Replace with your .fif project folder path
    HASH_VALUE = 12345  # You can change this value as needed
    OUTPUT_EXCEL = "fif_functions_output.xlsx"  # Name of the output Excel file

    found_functions = run_fif_decomposer(FOLDER_PATH, HASH_VALUE, OUTPUT_EXCEL)