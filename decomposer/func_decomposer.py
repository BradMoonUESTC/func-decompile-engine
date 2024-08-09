import os
import re
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
import time
import time
import re
import re

def extract_function_name(line):
    # 匹配 "_ %functionName()" 模式
    match = re.search(r'_\s*%(\w+)\s*\(', line)
    if match:
        return match.group(1)
    
    # 匹配 "(returnType) functionName(params)" 模式
    match = re.search(r'\([^)]*\)\s*([$\w]+)\s*\(', line)
    if match:
        return match.group(1)
    
    # 匹配以 $ 或 _ 开头，到下一个 ( 之间的内容
    match = re.search(r'([$_][\w$]+)\s*\(', line)
    if match:
        return match.group(1)
    
    return None

def find_func_functions(text, filename, hash):
    print(f"Starting to process file: {filename}")
    start_time = time.time()
    
    print(f"File size: {len(text)} characters")
    
    lines = text.split('\n')
    functions = []
    current_function = None
    brace_count = 0
    
    print("Starting line-by-line parsing...")
    for i, line in enumerate(lines):
        if i % 1000 == 0:
            print(f"Processed {i} lines...")
        
        stripped_line = line.strip()
        
        # 检查是否是函数开始
        if not current_function:
            func_name = extract_function_name(stripped_line)
            if func_name:
                current_function = {
                    'type': 'FunctionDefinition',
                    'name': func_name,
                    'content': line,
                    'start_line': i + 1,
                    'contract_name': filename.replace('.code.fc', '').replace('.storage.fc','').replace('stdlib.fc','').replace('constants.fc','').replace('native.fc','').replace('.',''),
                }
                brace_count = stripped_line.count('{')
        elif current_function:
            current_function['content'] += '\n' + line
            brace_count += line.count('{') - line.count('}')
            
            if brace_count == 0:
                # 函数结束
                current_function['end_line'] = i + 1
                functions.append(current_function)
                current_function = None
    
    end_time = time.time()
    print(f"Finished processing file: {filename}")
    print(f"Time taken: {end_time - start_time:.2f} seconds")
    print(f"Found {len(functions)} functions")

    return functions

def process_func_file(file_path, hash_value):
    with open(file_path, 'r', encoding='utf-8') as file:
        func_code = file.read()
    
    filename = os.path.basename(file_path)
    functions = find_func_functions(func_code, filename, hash_value)
    
    return functions

def process_func_folder(folder_path, hash_value):
    all_functions = []
    
    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith('.fc'):
                file_path = os.path.join(root, file)
                print(f"Processing file: {file_path}")
                functions = process_func_file(file_path, hash_value)
                all_functions.extend(functions)
                
                print(f"\nParsing results for {file_path}:")
                for func in functions:
                    print(json.dumps(func, indent=2))
                    print("-" * 50)
                
                print(f"Found {len(functions)} functions in {file_path}.")
    
    return all_functions

def test_find_func_functions(folder_path, hash_value, output_excel):
    if not os.path.isdir(folder_path):
        print(f"Error: '{folder_path}' is not a valid folder path.")
        return

    print(f"Processing folder: {folder_path}")
    all_functions = process_func_folder(folder_path, hash_value)

    print(f"\nTotal functions found: {len(all_functions)}.")

    # Filter out functions with name 'method_id' or 'if'
    filtered_functions = [func for func in all_functions if func['name'] not in ['method_id', 'if']]

    print(f"Functions after filtering: {len(filtered_functions)}")

    # Write results to Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FunC Functions"

    headers = ['Name', 'Start Line', 'End Line', 'Content', 'Contract Name', 'Return Type', 'Modifiers', 'Visibility', 'Node Count']
    ws.append(headers)

    # Style the header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for func in filtered_functions:
        ws.append([
            func['name'],
            func['start_line'],
            func['end_line'],
            func['content'],
            func['contract_name'],
            '',
            '',
            '',
            ''
        ])

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_excel)

    print(f"Results have been written to {output_excel}")

    return filtered_functions

def run_func_decomposer(folder_path, hash_value, output_excel):
    if not os.path.isdir(folder_path):
        print(f"Error: '{folder_path}' is not a valid folder path.")
        return None

    print(f"Processing folder: {folder_path}")
    all_functions = process_func_folder(folder_path, hash_value)

    print(f"\nTotal functions found: {len(all_functions)}.")

    # Filter out functions with name 'method_id' or 'if'
    filtered_functions = [func for func in all_functions if func['name'] not in ['method_id', 'if']]

    print(f"Functions after filtering: {len(filtered_functions)}")

    # Write results to Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FunC Functions"

    headers = ['Name', 'Start Line', 'End Line', 'Content', 'Contract Name', 'Return Type', 'Modifiers', 'Visibility', 'Node Count']
    ws.append(headers)

    # Style the header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for func in filtered_functions:
        ws.append([
            func['name'],
            func.get('start_line', ''),
            func.get('end_line', ''),
            func['content'],
            func['contract_name'],
            '',  # Return Type
            '',  # Modifiers
            '',  # Visibility
            ''   # Node Count
        ])

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_excel)

    print(f"Results have been written to {output_excel}")

    return filtered_functions

if __name__ == "__main__":
    FOLDER_PATH = "test/DexRouter"  # Replace with your FunC project folder path
    HASH_VALUE = 12345  # You can change this value as needed
    OUTPUT_EXCEL = "func_functions_output.xlsx"  # Name of the output Excel file

    found_functions = run_func_decomposer(FOLDER_PATH, HASH_VALUE, OUTPUT_EXCEL)